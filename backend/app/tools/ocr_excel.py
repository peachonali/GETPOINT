"""แปลงผลอ่านใบเสร็จหลายใบ → ไฟล์ Excel (.xlsx) สำหรับเก็บ/ตรวจ/ทำเฉลย

★ ต่างจาก send_queue/excel_export (ที่ทำไฟล์ให้อัปโหลดกลับ loga):
  ตัวนี้เป็น "ไฟล์สำหรับคนดู" — 1 แถวต่อ 1 ใบ พร้อมข้อความ OCR ดิบ
  ไว้ให้ผู้ใช้เอารูปมาป้อนเรื่อยๆ แล้วดูว่าระบบอ่านแม่นแค่ไหน + สะสมเป็นเฉลย
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

#: หัวตาราง (ไทย) — เรียงจากที่ใช้บ่อยไปข้อมูลดิบ
_COLUMNS = [
    ("filename", "ไฟล์", 22),
    ("ok", "อ่านได้", 8),
    ("total_amount", "ยอดเงิน", 12),
    ("merchant", "ร้าน", 26),
    ("merchant_code", "รหัสร้าน", 14),
    ("receipt_date", "วันที่", 12),
    ("receipt_time", "เวลา", 8),
    ("reference_codes", "เลขอ้างอิง", 24),
    ("items", "รายการสินค้า", 40),
    ("reason", "หมายเหตุ (ถ้าอ่านไม่ได้)", 26),
    ("raw_text", "ข้อความ OCR ดิบ", 50),
]


def build_excel(rows: list[dict[str, Any]]) -> bytes:
    """สร้างไฟล์ .xlsx จากผลอ่านหลายใบ · คืนเป็น bytes (ให้ดาวน์โหลด)"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "receipts"

    header_font = Font(bold=True)
    sheet.append([label for _key, label, _width in _COLUMNS])
    for cell in sheet[1]:
        cell.font = header_font

    for row in rows:
        sheet.append([_cell(row, key) for key, _label, _width in _COLUMNS])

    _apply_widths(sheet)
    _wrap_raw_text_column(sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _cell(row: dict[str, Any], key: str) -> Any:
    value = row.get(key)
    if key == "ok":
        return "✓" if value else "✗"
    if key == "total_amount":
        return "" if value is None else float(value)
    return value if value is not None else ""


def _apply_widths(sheet) -> None:
    for index, (_key, _label, width) in enumerate(_COLUMNS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width


def _wrap_raw_text_column(sheet) -> None:
    """ให้ช่อง "ข้อความ OCR ดิบ" ตัดบรรทัดในเซลล์ — จะได้เห็นหลายบรรทัดไม่ล้น"""
    raw_col = next(i for i, (k, _l, _w) in enumerate(_COLUMNS, start=1) if k == "raw_text")
    for row_cells in sheet.iter_rows(min_row=2, min_col=raw_col, max_col=raw_col):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
