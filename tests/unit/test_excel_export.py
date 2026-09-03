"""เทส app/send_queue/excel_export.py — ทางออก disaster recovery

★ กฎที่ต้องพิสูจน์:
    - เอาออกมาเฉพาะใบที่ "ยังไม่ได้แต้ม" (FAILED/PENDING) ไม่ใช่ใบที่ได้แต้มแล้ว
    - ต้องมี customer_id (cuid) ในไฟล์ ไม่งั้นอัปโหลดเข้า loga ไม่ได้ = กู้คืนไม่ได้จริง
    - คนละแบรนด์ต้องไม่ปนกัน
"""
import io

import pytest
from openpyxl import load_workbook

from app.database.members import Member
from app.database.receipts import (
    STATUS_AWARDED, STATUS_FAILED, STATUS_PENDING, STATUS_REJECTED, ReceiptRecord,
)
from app.database.tenants import Tenant
from app.send_queue.excel_export import export_unsent

TENANT = "v-club"
FORMULA = "7"


@pytest.fixture
def member_id(db_session) -> int:
    db_session.add(Tenant(id=TENANT, name="V-CLUB"))
    member = Member(tenant_id=TENANT, line_user_id="U1", crm_customer_id="CUST-1")
    db_session.add(member)
    db_session.commit()
    return member.id


def _add(session, member_id, *, amount, status, reference=None, tenant=TENANT):
    record = ReceiptRecord(
        tenant_id=tenant,
        member_id=member_id,
        content_fingerprint=f"fp-{amount}-{status}",
        image_fingerprint="img",
        merchant="ร้านทดสอบ",
        total_amount=amount,
        reference_codes=[],
        status=status,
        crm_reference=reference,
        source_image_id="img-1",
    )
    session.add(record)
    session.commit()
    return record


def _read(result_bytes):
    """อ่าน xlsx กลับมาเป็น list ของ dict (แถวแรกเป็นหัวตาราง)"""
    wb = load_workbook(io.BytesIO(result_bytes))
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


def test_exports_only_unsent_receipts(db_session, member_id):
    """★ ใบที่ได้แต้มแล้ว/ถูกปฏิเสธ ต้องไม่โผล่ในไฟล์กู้คืน"""
    _add(db_session, member_id, amount=100.0, status=STATUS_FAILED, reference="gp1")
    _add(db_session, member_id, amount=200.0, status=STATUS_PENDING, reference="gp2")
    _add(db_session, member_id, amount=300.0, status=STATUS_AWARDED, reference="gp3")
    _add(db_session, member_id, amount=400.0, status=STATUS_REJECTED, reference="gp4")

    result = export_unsent(db_session, TENANT, formula_id=FORMULA)

    assert result.row_count == 2
    refs = {row["reference"] for row in _read(result.content)}
    assert refs == {"gp1", "gp2"}


def test_export_includes_customer_id_for_reimport(db_session, member_id):
    """★★ ถ้าไม่มี cuid ในไฟล์ คนอัปโหลดเข้า loga ไม่ได้ = ไฟล์กู้คืนที่กู้ไม่ได้จริง"""
    _add(db_session, member_id, amount=149.0, status=STATUS_FAILED, reference="gp1")

    row = _read(export_unsent(db_session, TENANT, formula_id=FORMULA).content)[0]

    assert row["customer_id"] == "CUST-1"
    assert row["cost"] == "149.00"
    assert row["formula_id"] == FORMULA


def test_other_tenant_not_included(db_session, member_id):
    """คนละแบรนด์ต้องไม่ปนกันในไฟล์กู้คืน"""
    other = Member(tenant_id="other", line_user_id="U2", crm_customer_id="CUST-2")
    db_session.add(Tenant(id="other", name="Other"))
    db_session.add(other)
    db_session.commit()

    _add(db_session, member_id, amount=100.0, status=STATUS_FAILED, reference="mine")
    _add(db_session, other.id, amount=100.0, status=STATUS_FAILED, reference="theirs", tenant="other")

    refs = {row["reference"] for row in _read(export_unsent(db_session, TENANT, formula_id=FORMULA).content)}
    assert refs == {"mine"}


def test_empty_when_nothing_unsent(db_session, member_id):
    """ไม่มีใบค้าง → ไฟล์มีแต่หัวตาราง (ยังเปิดได้ ไม่พัง)"""
    _add(db_session, member_id, amount=100.0, status=STATUS_AWARDED, reference="gp1")

    result = export_unsent(db_session, TENANT, formula_id=FORMULA)
    assert result.row_count == 0
    assert _read(result.content) == []


def test_oldest_first(db_session, member_id):
    """เรียงเก่าสุดก่อน — คนกู้ทำตามลำดับที่เกิดจริง"""
    a = _add(db_session, member_id, amount=10.0, status=STATUS_FAILED, reference="first")
    b = _add(db_session, member_id, amount=20.0, status=STATUS_FAILED, reference="second")
    assert a.id < b.id

    refs = [row["reference"] for row in _read(export_unsent(db_session, TENANT, formula_id=FORMULA).content)]
    assert refs == ["first", "second"]
