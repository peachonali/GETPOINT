"""เทส /admin/* — หน้าจัดการหลังบ้าน

★ ประเด็นความปลอดภัยหลัก: ต้องกั้นด้วยโทเคนแอดมิน — ลูกค้าทั่วไปเข้าไม่ได้
  (dead-letter/export เผยข้อมูลการเงินรวม + ปลุกใบส่งซ้ำได้)
"""
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.database.db import get_session
from app.database.members import Member
from app.database.receipts import STATUS_AWARDED, STATUS_DEAD, STATUS_FAILED, ReceiptRecord
from app.database.tenants import Tenant
from app.main import app
from app.routes.dependencies import get_admin_token, get_formula_id, get_tenant_id

TENANT = "v-club"
TOKEN = "secret-admin-token"
GOOD = {"Authorization": f"Bearer {TOKEN}"}


@pytest.fixture
def ctx(db_session):
    db_session.add(Tenant(id=TENANT, name="V-CLUB"))
    member = Member(tenant_id=TENANT, line_user_id="U1", crm_customer_id="C1")
    db_session.add(member)
    db_session.commit()

    app.dependency_overrides[get_session] = lambda: db_session
    app.dependency_overrides[get_tenant_id] = lambda: TENANT
    app.dependency_overrides[get_admin_token] = lambda: TOKEN
    app.dependency_overrides[get_formula_id] = lambda: "7"

    yield {"client": TestClient(app), "session": db_session, "member_id": member.id}

    app.dependency_overrides.clear()


def _add(session, member_id, *, status, reference, amount=100.0, attempts=0):
    r = ReceiptRecord(
        tenant_id=TENANT, member_id=member_id,
        content_fingerprint=f"fp-{reference}", image_fingerprint="img",
        merchant="ร้าน", total_amount=amount, reference_codes=[],
        status=status, crm_reference=reference, send_attempts=attempts,
        source_image_id="img",
    )
    session.add(r)
    session.commit()
    return r


# ═══════════════════════════════════════════
# ★ ความปลอดภัย — ต้องมีโทเคนถูกต้อง
# ═══════════════════════════════════════════

def test_rejects_without_token(ctx):
    assert ctx["client"].get("/admin/metrics").status_code == 401


def test_rejects_wrong_token(ctx):
    resp = ctx["client"].get("/admin/metrics", headers={"Authorization": "Bearer wrong"})
    assert resp.status_code == 401


def test_disabled_when_no_token_configured(ctx):
    """★ ไม่ได้ตั้งโทเคน = ปิดหน้า admin ทั้งหมด (ปลอดภัยโดยปริยาย)"""
    app.dependency_overrides[get_admin_token] = lambda: ""
    assert ctx["client"].get("/admin/metrics", headers=GOOD).status_code == 401


def test_empty_token_not_bypassed_by_empty_bearer(ctx):
    """★★ ช่องโหว่คลาสสิก: ตั้งโทเคนว่าง + ส่ง Bearer ว่าง ต้องไม่หลุดเข้า

    ถ้าเช็ค "ปิดโดยปริยาย" หายไป compare_digest('', '') จะ match แล้วใครส่ง
    token ว่างก็เข้าได้หมด — ต้องกันด้วยการปฏิเสธ configured ว่างก่อนเทียบ
    """
    app.dependency_overrides[get_admin_token] = lambda: ""
    resp = ctx["client"].get("/admin/metrics", headers={"Authorization": "Bearer "})
    assert resp.status_code == 401


# ═══════════════════════════════════════════
# metrics
# ═══════════════════════════════════════════

def test_metrics(ctx):
    _add(ctx["session"], ctx["member_id"], status=STATUS_AWARDED, reference="a1")
    _add(ctx["session"], ctx["member_id"], status=STATUS_DEAD, reference="d1")

    body = ctx["client"].get("/admin/metrics", headers=GOOD).json()

    assert body["awarded"] == 1
    assert body["dead"] == 1
    assert body["needs_attention"] is True


# ═══════════════════════════════════════════
# dead letter
# ═══════════════════════════════════════════

def test_dead_letter_list(ctx):
    _add(ctx["session"], ctx["member_id"], status=STATUS_DEAD, reference="d1", attempts=5)
    _add(ctx["session"], ctx["member_id"], status=STATUS_FAILED, reference="f1")

    body = ctx["client"].get("/admin/dead-letter", headers=GOOD).json()

    assert body["dead_count"] == 1
    assert body["still_retrying_count"] == 1
    assert body["receipts"][0]["reference"] == "d1"
    assert body["receipts"][0]["attempts"] == 5


def test_revive_dead_letter(ctx):
    """★ ปลุกใบ DEAD กลับเข้าคิว → status กลับเป็น FAILED"""
    dead = _add(ctx["session"], ctx["member_id"], status=STATUS_DEAD, reference="d1", attempts=5)

    resp = ctx["client"].post(f"/admin/dead-letter/{dead.id}/revive", headers=GOOD)

    assert resp.json()["revived"] is True
    ctx["session"].refresh(dead)
    assert dead.status == STATUS_FAILED
    assert dead.send_attempts == 0


def test_revive_refuses_awarded(ctx):
    """★ ปลุกใบที่ได้แต้มแล้วไม่ได้ (จะให้แต้มซ้ำ)"""
    awarded = _add(ctx["session"], ctx["member_id"], status=STATUS_AWARDED, reference="a1")
    resp = ctx["client"].post(f"/admin/dead-letter/{awarded.id}/revive", headers=GOOD)
    assert resp.json()["revived"] is False


# ═══════════════════════════════════════════
# Excel export
# ═══════════════════════════════════════════

def test_export_downloads_xlsx(ctx):
    _add(ctx["session"], ctx["member_id"], status=STATUS_FAILED, reference="f1", amount=149.0)

    resp = ctx["client"].get("/admin/export.xlsx", headers=GOOD)

    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0][0] == "reference"          # หัวตาราง
    assert any("f1" in str(cell) for row in rows[1:] for cell in row)
