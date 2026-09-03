"""เทสเครื่องมือฝึก OCR — Excel builder + failed row + API routing

★ ไม่โหลด PaddleOCR จริงในเทส (ช้า/หนัก) — mock extract_one
  พิสูจน์ "โครงเครื่องมือ" ถูก: อัปโหลด → คืนแถว → ทำ Excel · ตัว OCR จริงวัดด้วย
  เครื่องมือ measure_* อยู่แล้ว
"""
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image

from app.tools import ocr_tool_api
from app.tools.ocr_excel import build_excel

JPEG = None


def _jpeg() -> bytes:
    buf = io.BytesIO()
    Image.new("RGB", (400, 600), (250, 250, 250)).save(buf, format="JPEG")
    return buf.getvalue()


# ═══════════════════════════════════════════
# Excel builder
# ═══════════════════════════════════════════

def test_excel_has_header_and_one_row_per_receipt():
    rows = [
        {"filename": "a.jpg", "ok": True, "reason": "", "merchant": "KFC", "merchant_code": "kfc",
         "total_amount": 149.0, "receipt_date": "2026-06-06", "receipt_time": "18:15",
         "reference_codes": "12102-002", "items": "BOX = 149", "raw_text": "KFC\nTotal 149"},
        {"filename": "b.jpg", "ok": False, "reason": "อ่านยอดไม่ได้", "merchant": "",
         "merchant_code": "", "total_amount": None, "receipt_date": "", "receipt_time": "",
         "reference_codes": "", "items": "", "raw_text": "เบลอมาก"},
    ]
    wb = load_workbook(io.BytesIO(build_excel(rows)))
    data = list(wb.active.iter_rows(values_only=True))

    assert data[0][0] == "ไฟล์"          # หัวตาราง
    assert len(data) == 3                 # หัว + 2 แถว
    assert data[1][2] == 149.0            # ยอดเงินเป็นตัวเลขจริง (คำนวณต่อได้)
    assert data[1][1] == "✓"
    assert data[2][1] == "✗"              # ใบที่อ่านไม่ได้


def test_excel_amount_blank_when_unreadable():
    rows = [{"filename": "x.jpg", "ok": False, "reason": "r", "merchant": "", "merchant_code": "",
             "total_amount": None, "receipt_date": "", "receipt_time": "", "reference_codes": "",
             "items": "", "raw_text": ""}]
    wb = load_workbook(io.BytesIO(build_excel(rows)))
    row = list(wb.active.iter_rows(values_only=True))[1]
    assert row[2] in ("", None)           # ยอดว่าง ไม่ใช่ 0 (0 จะทำให้เข้าใจผิดว่าฟรี)


# ═══════════════════════════════════════════
# API — mock ตัว OCR
# ═══════════════════════════════════════════

@pytest.fixture
def client(monkeypatch):
    """แทน extract_one ด้วยตัวปลอม — ไม่ต้องโหลดโมเดลจริง"""
    def fake_extract(filename, image):
        return {"filename": filename, "ok": True, "reason": "", "merchant": "KFC",
                "merchant_code": "kfc", "total_amount": 149.0, "receipt_date": "2026-06-06",
                "receipt_time": "18:15", "reference_codes": "abc", "items": "BOX = 149",
                "raw_text": "KFC 149"}

    monkeypatch.setattr(ocr_tool_api, "extract_one", fake_extract)
    return TestClient(ocr_tool_api.app)


def test_home_page_loads(client):
    resp = client.get("/")
    assert resp.status_code == 200
    assert "อ่านใบเสร็จ" in resp.text
    # ★ ต้องเปิดกล้องบนมือถือได้ (สแกน ไม่ใช่แค่แนบรูป)
    assert 'capture="environment"' in resp.text


def test_extract_returns_rows(client):
    resp = client.post("/api/extract", files=[("files", ("r.jpg", _jpeg(), "image/jpeg"))])
    body = resp.json()
    assert body["count"] == 1
    assert body["rows"][0]["merchant"] == "KFC"


def test_export_returns_xlsx(client):
    resp = client.post("/api/export.xlsx", files=[("files", ("r.jpg", _jpeg(), "image/jpeg"))])
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    wb = load_workbook(io.BytesIO(resp.content))
    assert list(wb.active.iter_rows(values_only=True))[1][3] == "KFC"


def test_non_image_becomes_failed_row_not_crash(client):
    """★ อัปไฟล์ไม่ใช่รูป → แถว "อ่านไม่ได้" ไม่ใช่พังทั้งชุด"""
    resp = client.post("/api/extract", files=[("files", ("bad.txt", b"not an image", "text/plain"))])
    body = resp.json()
    assert body["count"] == 1
    assert body["rows"][0]["ok"] is False


def test_rejects_empty_upload(client):
    resp = client.post("/api/extract", files=[])
    assert resp.status_code in (400, 422)
